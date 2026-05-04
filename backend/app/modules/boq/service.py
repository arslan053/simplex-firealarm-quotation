import logging
import uuid
from io import BytesIO

from fastapi import HTTPException, UploadFile, status
from sqlalchemy.ext.asyncio import AsyncSession

from app.modules.boq.images_handler import (
    build_combined_pdf,
    read_image_files,
    validate_image_filenames,
)
from app.modules.boq.pdf_handler import (
    upload_pdf_to_minio,
    validate_pdf_filename,
)
from app.modules.boq.repository import BoqItemRepository, DocumentRepository
from app.modules.boq.schemas import DocumentResponse
from app.shared.storage import upload_file
from app.shared.storage import delete_file
from app.shared.upload_security import (
    check_zip_bomb,
    sanitize_filename,
    sanitize_pdf,
    validate_file_size,
    validate_magic_bytes,
)

logger = logging.getLogger(__name__)

_XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_ALLOWED_EXCEL_EXTS = (".xlsx", ".xls")


def _convert_xls_to_xlsx(xls_bytes: bytes) -> bytes:
    """Convert legacy .xls bytes to .xlsx bytes using xlrd + openpyxl."""
    import xlrd
    from openpyxl import Workbook

    xls_book = xlrd.open_workbook(file_contents=xls_bytes)
    wb = Workbook()

    for idx, sheet_name in enumerate(xls_book.sheet_names()):
        xls_sheet = xls_book.sheet_by_index(idx)
        ws = wb.active if idx == 0 else wb.create_sheet()
        ws.title = sheet_name

        for row_idx in range(xls_sheet.nrows):
            for col_idx in range(xls_sheet.ncols):
                ws.cell(row=row_idx + 1, column=col_idx + 1,
                        value=xls_sheet.cell_value(row_idx, col_idx))

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class BoqService:
    def __init__(self, db: AsyncSession):
        self.db = db
        self.doc_repo = DocumentRepository(db)
        self.item_repo = BoqItemRepository(db)

    async def _delete_boq_document(
        self,
        tenant_id: uuid.UUID,
        doc,
    ) -> None:
        await self.item_repo.delete_by_document_id(doc.id, tenant_id)
        try:
            delete_file(doc.object_key)
        except Exception:
            logger.warning("Failed to delete BOQ file from MinIO: %s", doc.object_key)
        await self.doc_repo.delete(doc)

    async def _ensure_no_existing_boq_documents(
        self,
        tenant_id: uuid.UUID,
        project_id: uuid.UUID,
    ) -> None:
        docs = await self.doc_repo.get_by_project(tenant_id, project_id)
        if docs:
            raise HTTPException(
                status_code=status.HTTP_409_CONFLICT,
                detail="A BOQ document already exists. Remove it before uploading another BOQ.",
            )

    async def remove_boq_document(
        self,
        tenant_id: uuid.UUID,
        project_id: uuid.UUID,
        document_id: uuid.UUID,
    ) -> None:
        doc = await self.doc_repo.get_by_id(document_id, tenant_id, project_id)
        if not doc or doc.type != "BOQ":
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="BOQ document not found.",
            )
        await self._delete_boq_document(tenant_id, doc)

    # ------------------------------------------------------------------
    # Excel upload — store file only, no GPT extraction
    # ------------------------------------------------------------------

    async def upload_and_parse_boq(
        self,
        tenant_id: uuid.UUID,
        project_id: uuid.UUID,
        user_id: uuid.UUID,
        file: UploadFile,
    ) -> DocumentResponse:
        raw_name = file.filename or "unknown.xlsx"
        if not raw_name.lower().endswith(_ALLOWED_EXCEL_EXTS):
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Only .xlsx and .xls files are supported",
            )
        filename = sanitize_filename(raw_name)

        await self._ensure_no_existing_boq_documents(tenant_id, project_id)

        file_bytes = await file.read()
        validate_file_size(file_bytes)

        # Detect file type from extension for magic byte check
        ext = filename.rsplit(".", 1)[-1].lower()
        validate_magic_bytes(file_bytes, ext)

        # Convert legacy .xls → .xlsx so downstream parser (openpyxl) works
        if filename.lower().endswith(".xls") and not filename.lower().endswith(".xlsx"):
            logger.info("Converting .xls to .xlsx: %s", filename)
            file_bytes = _convert_xls_to_xlsx(file_bytes)
            filename = filename.rsplit(".", 1)[0] + ".xlsx"

        # ZIP bomb check on the final .xlsx bytes
        check_zip_bomb(file_bytes)

        file_size = len(file_bytes)

        file_uuid = uuid.uuid4()
        object_key = f"{tenant_id}/{project_id}/boq/{file_uuid}_{filename}"
        upload_file(
            object_key=object_key,
            data=file_bytes,
            content_type=_XLSX_MIME,
        )

        doc = await self.doc_repo.create(
            tenant_id=tenant_id,
            project_id=project_id,
            uploaded_by_user_id=user_id,
            original_file_name=filename,
            file_size=file_size,
            object_key=object_key,
        )
        return DocumentResponse.model_validate(doc)

    # ------------------------------------------------------------------
    # PDF upload — store file only, no GPT extraction
    # ------------------------------------------------------------------

    async def upload_and_parse_boq_pdf(
        self,
        tenant_id: uuid.UUID,
        project_id: uuid.UUID,
        user_id: uuid.UUID,
        file: UploadFile,
    ) -> DocumentResponse:
        raw_name = file.filename or "unknown.pdf"
        validate_pdf_filename(raw_name)
        filename = sanitize_filename(raw_name)

        await self._ensure_no_existing_boq_documents(tenant_id, project_id)

        pdf_bytes = await file.read()
        validate_file_size(pdf_bytes)
        validate_magic_bytes(pdf_bytes, "pdf")
        pdf_bytes = sanitize_pdf(pdf_bytes)
        file_size = len(pdf_bytes)

        object_key = upload_pdf_to_minio(tenant_id, project_id, filename, pdf_bytes)

        doc = await self.doc_repo.create(
            tenant_id=tenant_id,
            project_id=project_id,
            uploaded_by_user_id=user_id,
            original_file_name=filename,
            file_size=file_size,
            object_key=object_key,
        )
        return DocumentResponse.model_validate(doc)

    # ------------------------------------------------------------------
    # Images upload — combine into PDF and store, no GPT extraction
    # ------------------------------------------------------------------

    async def upload_and_parse_boq_images(
        self,
        tenant_id: uuid.UUID,
        project_id: uuid.UUID,
        user_id: uuid.UUID,
        files: list[UploadFile],
    ) -> DocumentResponse:
        validate_image_filenames(files)

        await self._ensure_no_existing_boq_documents(tenant_id, project_id)

        images = await read_image_files(files)

        pdf_bytes, filename = build_combined_pdf(images)
        file_size = len(pdf_bytes)

        file_uuid = uuid.uuid4()
        object_key = f"{tenant_id}/{project_id}/boq/{file_uuid}_{filename}"
        upload_file(
            object_key=object_key,
            data=pdf_bytes,
            content_type="application/pdf",
        )

        doc = await self.doc_repo.create(
            tenant_id=tenant_id,
            project_id=project_id,
            uploaded_by_user_id=user_id,
            original_file_name=filename,
            file_size=file_size,
            object_key=object_key,
        )
        return DocumentResponse.model_validate(doc)

    async def list_documents(
        self,
        tenant_id: uuid.UUID,
        project_id: uuid.UUID,
    ) -> list[DocumentResponse]:
        docs = await self.doc_repo.get_by_project(tenant_id, project_id)
        return [DocumentResponse.model_validate(doc) for doc in docs]
