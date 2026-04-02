from dataclasses import dataclass, field
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook


# Exact-match keyword sets (checked after lowercase + strip)
DESCRIPTION_EXACT = {"description", "desc", "item description", "part description", "parts description"}
QUANTITY_EXACT = {"qty", "qty.", "quantity"}
UNIT_EXACT = {"unit", "uom"}
TYPE_EXACT = {"type", "item_type"}

# Substring keywords — if any of these appear inside the header, it matches
DESCRIPTION_SUBSTRINGS = ["desc"]
QUANTITY_SUBSTRINGS = ["qty", "quantity"]
UNIT_SUBSTRINGS = ["unit", "uom"]
TYPE_SUBSTRINGS = ["type"]

_ALLOWED_TYPES = {"boq_item", "description", "section_description"}


@dataclass
class ParsedItem:
    row_number: int
    description: str | None
    quantity: Decimal | None
    unit: str | None
    is_valid: bool
    type: str = "boq_item"


@dataclass
class ParseResult:
    success: bool
    items: list[ParsedItem] = field(default_factory=list)
    missing_columns: list[str] = field(default_factory=list)


def _normalize_header(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().lower()


def _parse_quantity(value: object) -> Decimal | None:
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _match_column(header: str, exact_set: set[str], substrings: list[str]) -> bool:
    """Two-tier matching: exact match first, then substring match."""
    if header in exact_set:
        return True
    for kw in substrings:
        if kw in header:
            return True
    return False


def _is_any_known_header(header: str) -> bool:
    """Check if a header matches any known column type (for header row detection)."""
    return (
        _match_column(header, DESCRIPTION_EXACT, DESCRIPTION_SUBSTRINGS)
        or _match_column(header, QUANTITY_EXACT, QUANTITY_SUBSTRINGS)
        or _match_column(header, UNIT_EXACT, UNIT_SUBSTRINGS)
        or _match_column(header, TYPE_EXACT, TYPE_SUBSTRINGS)
    )


def parse_boq_excel(file_bytes: bytes) -> ParseResult:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active

    # Find header row — scan first 10 rows
    header_row_idx = None
    headers: dict[int, str] = {}  # col_idx -> normalized header name

    for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=False), start=1):
        normalized = [_normalize_header(cell.value) for cell in row]
        # Check if this row contains at least 2 known header variants
        matches = sum(1 for val in normalized if val and _is_any_known_header(val))
        if matches >= 2:
            header_row_idx = row_idx
            for col_idx, val in enumerate(normalized):
                headers[col_idx] = val
            break

    if header_row_idx is None:
        # Fallback: use first row
        header_row_idx = 1
        first_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=False))
        for col_idx, cell in enumerate(first_row):
            headers[col_idx] = _normalize_header(cell.value)

    # Map columns using two-tier matching
    desc_col = None
    qty_col = None
    unit_col = None
    type_col = None

    for col_idx, header in headers.items():
        if not header:
            continue
        # Check type first to avoid "type" matching description substrings
        if type_col is None and _match_column(header, TYPE_EXACT, TYPE_SUBSTRINGS):
            type_col = col_idx
        elif desc_col is None and _match_column(header, DESCRIPTION_EXACT, DESCRIPTION_SUBSTRINGS):
            desc_col = col_idx
        elif qty_col is None and _match_column(header, QUANTITY_EXACT, QUANTITY_SUBSTRINGS):
            qty_col = col_idx
        elif unit_col is None and _match_column(header, UNIT_EXACT, UNIT_SUBSTRINGS):
            unit_col = col_idx

    # Check required columns (only description and quantity are required)
    missing = []
    if desc_col is None:
        missing.append("description")
    if qty_col is None:
        missing.append("quantity")

    if missing:
        wb.close()
        return ParseResult(success=False, missing_columns=missing)

    # Parse data rows — row numbers are always sequential (1, 2, 3...)
    items: list[ParsedItem] = []
    seq = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        # Skip completely blank rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        seq += 1

        # Extract values
        desc_val = str(row[desc_col]).strip() if desc_col < len(row) and row[desc_col] is not None else None
        qty_val = _parse_quantity(row[qty_col]) if qty_col < len(row) else None
        unit_val = None
        if unit_col is not None:
            unit_val = str(row[unit_col]).strip() if unit_col < len(row) and row[unit_col] is not None else None

        # Read type value
        row_type = "boq_item"
        if type_col is not None and type_col < len(row) and row[type_col] is not None:
            raw_type = str(row[type_col]).strip().lower()
            if raw_type in _ALLOWED_TYPES:
                row_type = raw_type

        # Empty string → None
        if desc_val == "":
            desc_val = None
        if unit_val == "":
            unit_val = None

        # Determine validity
        # 'description' and 'section_description' rows are always valid
        # 'boq_item' rows require description and quantity
        if row_type in ("description", "section_description"):
            is_valid = True
        else:
            is_valid = desc_val is not None and qty_val is not None

        items.append(ParsedItem(
            row_number=seq,
            description=desc_val,
            quantity=qty_val,
            unit=unit_val,
            is_valid=is_valid,
            type=row_type,
        ))

    wb.close()
    return ParseResult(success=True, items=items)
