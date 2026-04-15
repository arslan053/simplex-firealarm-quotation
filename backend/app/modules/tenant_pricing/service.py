from __future__ import annotations

import io
import uuid

from fastapi import HTTPException
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Protection
from openpyxl.utils import get_column_letter
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession

from app.shared.upload_security import (
    check_zip_bomb,
    validate_file_size,
    validate_magic_bytes,
)

from .schemas import (
    PriceListResponse,
    PriceUpdateResponse,
    TemplateValidationError,
    TenantProductPrice,
    UploadResponse,
)


class TenantPricingService:
    def __init__(self, db: AsyncSession):
        self.db = db

    async def get_price_list(
        self,
        tenant_id: uuid.UUID,
        search: str | None = None,
        category: str | None = None,
    ) -> PriceListResponse:
        conditions = ["1=1"]
        params: dict = {"tid": tenant_id}

        if search:
            conditions.append(
                "(p.code ILIKE :search OR p.description ILIKE :search)"
            )
            params["search"] = f"%{search}%"

        if category:
            conditions.append("p.category::text = :category")
            params["category"] = category

        where = " AND ".join(conditions)

        result = await self.db.execute(
            text(f"""
                SELECT
                    p.id, p.code, p.description, p.category::text,
                    COALESCE(tpp.price, 0) AS price,
                    COALESCE(tpp.currency, p.currency) AS currency
                FROM products p
                LEFT JOIN tenant_product_prices tpp
                    ON tpp.product_id = p.id AND tpp.tenant_id = :tid
                WHERE {where}
                ORDER BY p.category::text, p.code
            """),
            params,
        )
        rows = result.fetchall()

        items = [
            TenantProductPrice(
                product_id=str(r[0]),
                code=r[1],
                description=r[2],
                category=r[3],
                price=float(r[4]),
                currency=r[5],
            )
            for r in rows
        ]
        prices_set = sum(1 for item in items if item.price > 0)

        return PriceListResponse(
            items=items,
            total=len(items),
            prices_set=prices_set,
        )

    async def update_prices(
        self,
        tenant_id: uuid.UUID,
        items: list[dict],
    ) -> PriceUpdateResponse:
        updated = 0
        for item in items:
            await self.db.execute(
                text("""
                    INSERT INTO tenant_product_prices (tenant_id, product_id, price, updated_at)
                    VALUES (:tid, CAST(:pid AS uuid), :price, now())
                    ON CONFLICT (tenant_id, product_id)
                    DO UPDATE SET price = EXCLUDED.price, updated_at = now()
                """),
                {"tid": tenant_id, "pid": item["product_id"], "price": item["price"]},
            )
            updated += 1
        await self.db.commit()
        return PriceUpdateResponse(updated=updated)

    async def get_categories(self) -> list[str]:
        result = await self.db.execute(
            text("SELECT DISTINCT category::text FROM products ORDER BY category::text")
        )
        return [r[0] for r in result.fetchall()]

    async def generate_template(self, tenant_id: uuid.UUID) -> bytes:
        result = await self.db.execute(
            text("""
                SELECT
                    p.id, p.code, p.description,
                    COALESCE(tpp.price, 0) AS price
                FROM products p
                LEFT JOIN tenant_product_prices tpp
                    ON tpp.product_id = p.id AND tpp.tenant_id = :tid
                ORDER BY p.category::text, p.code
            """),
            {"tid": tenant_id},
        )
        rows = result.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Price List"

        # Styles
        header_font = Font(bold=True, size=11)
        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        locked = Protection(locked=True)
        unlocked = Protection(locked=False)
        locked_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

        # Headers
        headers = ["Product ID", "Code", "Description", "Price (USD)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.protection = locked

        # Data rows
        for i, row in enumerate(rows, 2):
            # Product ID (locked)
            c_id = ws.cell(row=i, column=1, value=str(row[0]))
            c_id.protection = locked
            c_id.fill = locked_fill

            # Code (locked)
            c_code = ws.cell(row=i, column=2, value=row[1])
            c_code.protection = locked
            c_code.fill = locked_fill

            # Description (locked)
            c_desc = ws.cell(row=i, column=3, value=row[2])
            c_desc.protection = locked
            c_desc.fill = locked_fill

            # Price (editable)
            c_price = ws.cell(row=i, column=4, value=float(row[3]))
            c_price.protection = unlocked
            c_price.number_format = '#,##0.000'

        # Column widths
        ws.column_dimensions['A'].width = 38
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 45
        ws.column_dimensions['D'].width = 15

        # Enable sheet protection — locked cells can't be edited
        ws.protection.sheet = True
        ws.protection.enable()

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    async def process_upload(
        self,
        tenant_id: uuid.UUID,
        file_bytes: bytes,
    ) -> UploadResponse:
        validate_file_size(file_bytes)
        validate_magic_bytes(file_bytes, "xlsx")
        check_zip_bomb(file_bytes)

        try:
            wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
        except Exception:
            raise HTTPException(
                status_code=400,
                detail="The file is not a valid Excel spreadsheet (.xlsx).",
            )

        ws = wb.active

        # Build a lookup of all products for validation
        result = await self.db.execute(
            text("SELECT id, code, description FROM products")
        )
        product_map: dict[str, dict] = {}
        for r in result.fetchall():
            product_map[str(r[0])] = {"code": r[1], "description": r[2]}

        errors: list[TemplateValidationError] = []
        to_upsert: list[dict] = []

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 4:
                continue

            raw_id = str(row[0]).strip() if row[0] else ""
            raw_code = str(row[1]).strip() if row[1] else ""
            raw_desc = str(row[2]).strip() if row[2] else ""
            raw_price = row[3]

            # Skip rows with no price
            if raw_price is None or raw_price == "" or raw_price == 0:
                continue

            try:
                price = float(raw_price)
            except (ValueError, TypeError):
                errors.append(TemplateValidationError(
                    row=row_idx,
                    expected_code=raw_code,
                    got_code=raw_code,
                    message=f"Invalid price value: {raw_price}",
                ))
                continue

            # Validate product id + code + description match
            product = product_map.get(raw_id)
            if not product:
                errors.append(TemplateValidationError(
                    row=row_idx,
                    expected_code="(unknown)",
                    got_code=raw_code,
                    message=f"Product ID not found: {raw_id}",
                ))
                continue

            if product["code"] != raw_code:
                errors.append(TemplateValidationError(
                    row=row_idx,
                    expected_code=product["code"],
                    got_code=raw_code,
                    message=f"Code mismatch — expected '{product['code']}', got '{raw_code}'",
                ))
                continue

            if product["description"].strip() != raw_desc:
                errors.append(TemplateValidationError(
                    row=row_idx,
                    expected_code=product["code"],
                    got_code=raw_code,
                    message=f"Description mismatch for {raw_code} — file was modified",
                ))
                continue

            to_upsert.append({"product_id": raw_id, "price": price})

        # Batch upsert valid rows
        for item in to_upsert:
            await self.db.execute(
                text("""
                    INSERT INTO tenant_product_prices (tenant_id, product_id, price, updated_at)
                    VALUES (:tid, CAST(:pid AS uuid), :price, now())
                    ON CONFLICT (tenant_id, product_id)
                    DO UPDATE SET price = EXCLUDED.price, updated_at = now()
                """),
                {"tid": tenant_id, "pid": item["product_id"], "price": item["price"]},
            )
        if to_upsert:
            await self.db.commit()

        wb.close()
        return UploadResponse(updated=len(to_upsert), errors=errors)
