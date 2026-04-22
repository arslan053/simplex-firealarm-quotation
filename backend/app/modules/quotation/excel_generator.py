"""Quotation XLSX generator.

Builds a professional quotation spreadsheet that visually matches the DOCX output.
Uses openpyxl to produce an Excel file that LibreOffice can convert to PDF.

All spacing values are derived from the DOCX generator's twip-based paragraph
spacing to produce an identical layout.
"""
from __future__ import annotations

import io
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.pagebreak import Break

from .generator import QuotationData, _ordinal_suffix, _format_price, _format_qty

logger = logging.getLogger(__name__)

_TEMPLATES_DIR = Path(__file__).parent / "templates"
_IMAGES_DIR = _TEMPLATES_DIR / "images"

# ── Fonts (matching DOCX: Verdana 10pt body, 9pt tables) ──
_FONT_NAME = "Verdana"
_FONT_BODY = Font(name=_FONT_NAME, size=10)
_FONT_BODY_BOLD = Font(name=_FONT_NAME, size=10, bold=True)
_FONT_BODY_BOLD_UL = Font(name=_FONT_NAME, size=10, bold=True, underline="single")
_FONT_SMALL = Font(name=_FONT_NAME, size=9)
_FONT_SMALL_BOLD = Font(name=_FONT_NAME, size=9, bold=True)

# ── Alignments ──
_WRAP_TOP = Alignment(wrap_text=True, vertical="top")
_WRAP_LEFT = Alignment(horizontal="left", vertical="top", wrap_text=True)
_WRAP_RIGHT = Alignment(horizontal="right", vertical="top", wrap_text=True)
_WRAP_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# ── Borders ──
_NO_BORDER = Border(
    left=Side(style=None), right=Side(style=None),
    top=Side(style=None), bottom=Side(style=None),
)
_THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
_GRAY_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

# ── Column widths ──
# A=Model, B=Description, C=Qty, D=Unit Price, E=Total Price
# Mapped from DOCX Cm: 2.62 + 9.82 + 1.50 + 2.30 + 2.50 = 18.74cm
# In Excel char units (1 char ≈ 0.19cm for Verdana): scale each to match
_COL_WIDTHS = {"A": 14, "B": 52, "C": 8, "D": 12, "E": 13}

# ── DOCX paragraph spacing (twips) → Excel row heights (points) ──
# 200 twips = 10pt (section header space-before)
#  60 twips =  3pt (section header space-after)
# 120 twips =  6pt (body text space-after)
# 160 twips =  8pt (payment terms space-before)
_SP_SECTION_BEFORE = 10  # pt — gap row before section headers
_SP_BODY_AFTER = 3       # pt — tiny gap between body paragraphs

# Row height for one line of Verdana 10pt with a bit of breathing room
_LINE_HEIGHT = 14.5  # pt
_LINE_HEIGHT_SMALL = 13  # pt (for 9pt font in tables)

# ── Page margins (inches) ──
_TOP_MARGIN = 4.30 / 2.54
_BOTTOM_MARGIN = 2.38 / 2.54
_LEFT_MARGIN = 1.91 / 2.54
_RIGHT_MARGIN = 2.11 / 2.54

# Characters per line at merged A:E width for height estimation
_CHARS_PER_LINE = 95


def generate_quotation_xlsx(data: QuotationData) -> bytes:
    """Generate a quotation XLSX and return it as bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Quotation"

    # ── Make it look like a document, not a spreadsheet ──
    ws.sheet_view.showGridLines = False

    # ── Page setup (A4 portrait, matching DOCX margins) ──
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.top = _TOP_MARGIN
    ws.page_margins.bottom = _BOTTOM_MARGIN
    ws.page_margins.left = _LEFT_MARGIN
    ws.page_margins.right = _RIGHT_MARGIN
    ws.page_margins.header = 0.38
    ws.page_margins.footer = 0.80

    for col_letter, width in _COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

    # ── Extract letterhead images ──
    header_images, footer_images = _extract_letterhead_images(data.letterhead_bytes)

    row = 1

    # ═══════════════════════════════════════════════════════════
    # PAGE 1: Cover letter
    # ═══════════════════════════════════════════════════════════

    row = _add_letterhead(ws, row, header_images)
    row = _add_client_and_date(ws, row, data)

    # Subject (no extra gap — DOCX adds it right after address)
    row = _write_text(ws, row, f"Subject: {data.subject or f'Fire Alarm System – Simplex- {data.project_name}'}", _FONT_BODY)

    # Greeting (DOCX: just another paragraph, no blank row)
    row = _write_text(ws, row, f"Dear Engr, {data.client_name}", _FONT_BODY)

    # ── Service-option-dependent body ──
    if data.service_option == 1:
        row = _write_text(ws, row,
            "As requested, please find herewith attached our offer for Fire Alarm System.",
            _FONT_BODY)
        row = _write_section(ws, row, "SCOPE",
            "Price includes Supply of equipment mentioned in attached point-schedule, "
            "warranty, programming, testing & commissioning.")
        row = _write_section(ws, row, "WARRANTY:",
            "Items supplied by us shall be covered under our standard warranty clause that "
            "covers against any material defect or malfunctioning, for a period of 18 months "
            "from date of delivery. Product shall be used as intended. Misuse or wrong application "
            "will not be covered under warranty. We also hope that project maintenance will be "
            "given to us as a separate contract so that we can maintain the system in a proper way.")
        row = _write_text(ws, row,
            "However our warranty coverage shall not include wear & tear, consumables, "
            "abuse/ misuse/ wrong use of components",
            _FONT_BODY)
    else:
        row = _write_text(ws, row,
            "As requested, please find herewith attached our offer for above mentioned "
            "systems. The Offer is based on the drawings provided. We have proposed "
            "material supply and installation as requested.",
            _FONT_BODY)
        scope_text = (
            "Price includes Supply of equipment mentioned in attached point-schedule, "
            "engineering support which includes preparation of Single Line diagrams, "
            "Installation of devices we supplied, "
            + ("conduiting, " if data.service_option == 3 else "")
            + "cable pulling, device fixing, "
            "programming, testing and commissioning of equipment we supplied, Client Staff "
            "training, O&M Manuals and Warranty support."
        )
        row = _write_section(ws, row, "SCOPE", scope_text)
        row = _write_text(ws, row,
            "Any changes in architectural design or material list or "
            "system design will change price.",
            _FONT_BODY)
        row = _write_section(ws, row, "WARRANTY:",
            "However, our warranty coverage shall not include wear & tear, consumables, "
            "abuse/ misuse/ wrong use of components.")

    # ── Common sections (matching DOCX order exactly) ──
    row = _write_section(ws, row, "CANCELLATION:",
        "In case of cancellation of order for whatsoever reasons RGM reserves the right "
        "to charge the purchaser for the cost of such cancellations in accordance with "
        "the actual stage of processing.")

    row = _add_spacer(ws, row, _SP_SECTION_BEFORE)
    row = _write_text(ws, row, "LIMITATION OF LIABILITY:", _FONT_BODY_BOLD_UL)
    row = _write_text(ws, row,
        "The supplier shall not be liable, whether in contract, warranty, failure of remedy "
        "to achieve its essential purpose, tort (including negligence or strict liability) "
        "indemnity, or any other legal or equitable theory for damage to or loss of other "
        "property or equipment, business interruption or lost revenue, profits or sales, "
        "cost of capital, or for any special, incidental, punitive, indirect or consequential "
        "damages or for any other loss, costs or expenses of similar type.",
        _FONT_BODY)
    row = _write_text(ws, row,
        "The liability of the supplier for any act or omission, product sold, serviced or "
        "furnished directly or indirectly under this agreement, whether in contract, warranty "
        "failure or a remedy to achieve its essential purpose, tort (including negligence or "
        "strict liability) indemnity, or any other legal or equitable theory, will in no event "
        "exceed 1% of the contract value.",
        _FONT_BODY)
    row = _write_text(ws, row,
        "The rights and remedies contained in this agreement are exclusive, and the parties "
        "accept these remedies in lieu of all other rights and remedies available at law or "
        "otherwise, in contract (including warranty) or in tort (including negligence), for "
        "any and all claims of any nature arising under this agreement or any performance or "
        "breach arising out of this agreement.",
        _FONT_BODY)

    # NOTES & EXCLUSIONS
    row = _add_notes_exclusions(ws, row, data)

    # PRICES AND TERMS OF PAYMENT
    row = _add_spacer(ws, row, _SP_SECTION_BEFORE)
    row = _write_text(ws, row, "PRICES AND TERMS OF PAYMENT:", _FONT_BODY_BOLD_UL)
    row = _write_text(ws, row,
        "Without prejudice to any further rights, we may suspend and/ or refuse any supplies "
        "for as long as any due payment remains outstanding for whatsoever reason.",
        _FONT_BODY)
    row = _write_text(ws, row,
        "Late payments due and payable to supplier shall attract interest at a rated of 12% "
        "per annum accruing from their due date until full settlement of the principal amount. "
        "Payments by the purchaser shall be deemed to be made first against any accrued interest "
        "and then against the outstanding principal amount. The provision of this clause is "
        "without prejudice to any further rights of the supplier in case of payment is delayed "
        "by the purchaser.",
        _FONT_BODY)

    # Payment terms (DOCX: space_before=160 twips=8pt)
    row = _add_spacer(ws, row, 8)
    row = _write_text(ws, row, "Payment terms :", _FONT_BODY_BOLD)
    if data.payment_terms_text:
        for line in data.payment_terms_text.splitlines():
            if line.strip():
                row = _write_text(ws, row, line, _FONT_BODY)

    # TIME FOR SUPPLIES
    row = _write_section(ws, row, "TIME FOR SUPPLIES:",
        "Delivery \u2013 14 to 16 weeks from the date of advance payment receipt. "
        "(Partial Delivery against partial payment allowed)")

    # Validity + Signature (DOCX: space_before=200, space_after=100)
    row = _add_spacer(ws, row, _SP_SECTION_BEFORE)
    row = _write_text(ws, row, "Validity \u2013 30 days", _FONT_BODY_BOLD)
    if data.signatory_name:
        row = _write_text(ws, row, "Best regards,", _FONT_BODY)
        row = _write_text(ws, row, data.signatory_name, _FONT_BODY)
    if data.company_phone:
        row = _write_text(ws, row, data.company_phone, _FONT_BODY)
    if data.signature_bytes:
        row = _add_image_from_bytes(ws, row, data.signature_bytes, width_cm=3.5)

    # ═��═════════════════════════════════════════════════════════
    # PAGE 2: Product table (new page, matching DOCX page break)
    # ═══════════════════════════════════════════════════════════
    ws.row_breaks.append(Break(id=row - 1))
    row = _add_product_table(ws, row, data)

    # Footer image
    if footer_images:
        row += 1
        img = XlImage(io.BytesIO(footer_images[0]))
        if img.width and img.height:
            aspect = img.height / img.width
            img.width = 756
            img.height = int(img.width * aspect)
        ws.add_image(img, f"A{row}")
        row += 4
    else:
        footer_path = _IMAGES_DIR / "footer.png"
        if footer_path.exists():
            row += 1
            img = XlImage(str(footer_path))
            img.width = 756
            img.height = 86
            ws.add_image(img, f"A{row}")
            row += 4

    ws.print_area = f"A1:E{row + 2}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ═══════════════════════════════════════════════════════════════
# Core text helpers — replicate DOCX paragraph behavior
# ════════��═══════════════════════════════════════════���══════════

def _write_text(ws, row: int, text: str, font: Font) -> int:
    """Write a single paragraph as a merged A:E row. No borders.
    Returns the next available row.
    """
    ws.merge_cells(f"A{row}:E{row}")
    cell = ws[f"A{row}"]
    cell.value = text
    cell.font = font
    cell.alignment = _WRAP_TOP
    cell.border = _NO_BORDER

    # Height: count wrapped lines (Verdana 10pt at ~95 chars/line)
    n_lines = _count_lines(text, _CHARS_PER_LINE)
    ws.row_dimensions[row].height = n_lines * _LINE_HEIGHT
    return row + 1


def _write_section(ws, row: int, header: str, body: str) -> int:
    """Write a section header + body paragraph — matching DOCX
    _set_para_spacing(before=200, after=60) + _add_body_text().
    """
    row = _add_spacer(ws, row, _SP_SECTION_BEFORE)
    row = _write_text(ws, row, header, _FONT_BODY_BOLD_UL)
    row = _write_text(ws, row, body, _FONT_BODY)
    return row


def _add_spacer(ws, row: int, height_pt: float) -> int:
    """Insert a tiny spacer row (like paragraph space-before in DOCX)."""
    ws.row_dimensions[row].height = height_pt
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"].border = _NO_BORDER
    return row + 1


def _count_lines(text: str, chars_per_line: int) -> int:
    """Estimate number of wrapped lines for row height calculation."""
    lines = 0
    for paragraph in text.split("\n"):
        lines += max(1, -(-len(paragraph) // chars_per_line))  # ceiling division
    return lines


# ═══════════════════════════════════════════════════════════════
# Letterhead
# ══════════════════════════════════════════════════════���════════

def _extract_letterhead_images(letterhead_bytes: bytes | None) -> tuple[list[bytes], list[bytes]]:
    """Extract header and footer images from a letterhead DOCX.
    Falls back to on-disk images if no letterhead uploaded.
    """
    if not letterhead_bytes:
        header_imgs: list[bytes] = []
        for name in ("logo.png", "header_text.png"):
            p = _IMAGES_DIR / name
            if p.exists():
                header_imgs.append(p.read_bytes())
        footer_imgs: list[bytes] = []
        if (_IMAGES_DIR / "footer.png").exists():
            footer_imgs.append((_IMAGES_DIR / "footer.png").read_bytes())
        return header_imgs, footer_imgs

    try:
        from app.shared.upload_security import sanitize_docx_package
        from docx import Document

        clean_bytes = sanitize_docx_package(letterhead_bytes)
        doc = Document(io.BytesIO(clean_bytes))

        header_imgs = []
        footer_imgs = []
        for section in doc.sections:
            hdr = section.header
            if not hdr.is_linked_to_previous:
                for rel in hdr.part.rels.values():
                    if "image" in rel.reltype:
                        try:
                            header_imgs.append(rel.target_part.blob)
                        except Exception:
                            pass
            ftr = section.footer
            if not ftr.is_linked_to_previous:
                for rel in ftr.part.rels.values():
                    if "image" in rel.reltype:
                        try:
                            footer_imgs.append(rel.target_part.blob)
                        except Exception:
                            pass
        return header_imgs, footer_imgs
    except Exception:
        logger.warning("Failed to extract letterhead images from DOCX", exc_info=True)
        return [], []


def _add_letterhead(ws, row: int, header_images: list[bytes]) -> int:
    """Place letterhead images at top of sheet."""
    if not header_images:
        return row

    if len(header_images) >= 2:
        # Logo (left) + header text (right) — matching DOCX anchored positions
        logo = XlImage(io.BytesIO(header_images[0]))
        logo.width = 100
        logo.height = 105
        ws.add_image(logo, f"A{row}")

        htxt = XlImage(io.BytesIO(header_images[1]))
        htxt.width = 450
        htxt.height = 75
        ws.add_image(htxt, f"C{row}")
    else:
        # Single image — span across
        img = XlImage(io.BytesIO(header_images[0]))
        if img.width and img.height:
            aspect = img.height / img.width
            img.width = 600
            img.height = int(img.width * aspect)
        ws.add_image(img, f"A{row}")

    # Reserve 4 rows for header images
    for r in range(row, row + 4):
        ws.row_dimensions[r].height = 20
    row += 4

    # Gray separator line (matching DOCX 0.5pt A0A0A0 bottom border)
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"].border = Border(bottom=Side(style="thin", color="A0A0A0"))
    ws.row_dimensions[row].height = 4
    row += 1
    return row


# ══════���═══════════════════════���════════════════════════════════
# Client / Date block
# ════════════════════��══════════════════════════════════════════

def _add_client_and_date(ws, row: int, data: QuotationData) -> int:
    """Engr, <name>  [right] Date: 22nd Apr 2026
       <address>,     [right] Ref.: MI/203-C/...
    Matches DOCX right-tab-stop layout.
    """
    d = data.generation_date
    day_suffix = _ordinal_suffix(d.day)
    date_str = f"Date: {d.day}{day_suffix} {d.strftime('%b')} {d.year}"
    ref_str = f"Ref.: MI/203-C/{data.reference_number}"

    # Line 1
    ws.merge_cells(f"A{row}:C{row}")
    c = ws[f"A{row}"]
    c.value = f"Engr, {data.client_name}"
    c.font = _FONT_BODY_BOLD
    c.alignment = _WRAP_LEFT
    c.border = _NO_BORDER

    ws.merge_cells(f"D{row}:E{row}")
    c = ws[f"D{row}"]
    c.value = date_str
    c.font = _FONT_BODY_BOLD
    c.alignment = _WRAP_RIGHT
    c.border = _NO_BORDER
    ws.row_dimensions[row].height = _LINE_HEIGHT
    row += 1

    # Line 2
    ws.merge_cells(f"A{row}:C{row}")
    c = ws[f"A{row}"]
    c.value = f"{data.client_address},"
    c.font = _FONT_BODY
    c.alignment = _WRAP_LEFT
    c.border = _NO_BORDER

    ws.merge_cells(f"D{row}:E{row}")
    c = ws[f"D{row}"]
    c.value = ref_str
    c.font = _FONT_BODY_BOLD
    c.alignment = _WRAP_RIGHT
    c.border = _NO_BORDER
    ws.row_dimensions[row].height = _LINE_HEIGHT
    row += 1

    # Small gap (not a full blank row — matches DOCX default paragraph spacing)
    row = _add_spacer(ws, row, 6)
    return row


# ══════════════��════════════════════════════════════════════════
# Notes & Exclusions table
# ═══════════════════════════════════════════════════════════════

def _add_notes_exclusions(ws, row: int, data: QuotationData) -> int:
    """Bordered 2-column mini-table matching DOCX Table Grid style."""
    from .inclusions import build_document_items

    items = build_document_items(data.service_option, data.inclusion_answers or {})
    if not items:
        return row

    row = _add_spacer(ws, row, _SP_SECTION_BEFORE)
    row = _write_text(ws, row, "NOTES & EXCLUSIONS:", _FONT_BODY_BOLD_UL)

    for i, item_text in enumerate(items):
        # Number cell (narrow — column A only)
        num_cell = ws[f"A{row}"]
        num_cell.value = i + 1
        num_cell.font = _FONT_SMALL
        num_cell.alignment = _WRAP_CENTER
        num_cell.border = _THIN_BORDER

        # Text cell (columns B:E merged)
        ws.merge_cells(f"B{row}:E{row}")
        txt_cell = ws[f"B{row}"]
        txt_cell.value = item_text
        txt_cell.font = _FONT_SMALL
        txt_cell.alignment = Alignment(wrap_text=True, vertical="center")
        txt_cell.border = _THIN_BORDER
        for col in ("C", "D", "E"):
            ws[f"{col}{row}"].border = _THIN_BORDER

        # Height: ~80 chars/line for B:E width, 9pt font
        n_lines = _count_lines(item_text, 80)
        ws.row_dimensions[row].height = max(_LINE_HEIGHT_SMALL, n_lines * _LINE_HEIGHT_SMALL)
        row += 1

    return row


# ═══���══════════════════════════��════════════════════════════════
# Product table (page 2)
# ════════════��═══════════════════════════��══════════════════════

def _add_product_table(ws, row: int, data: QuotationData) -> int:
    """5-column product table matching DOCX layout exactly."""
    has_installation = data.installation_amount > 0

    # Header row (matching DOCX: " Unit\nPrice ", " Total\nPrice ")
    headers = ["Model", "Description", "Qty", " Unit\nPrice ", " Total\nPrice "]
    for ci, header in enumerate(headers):
        cell = ws.cell(row=row, column=ci + 1)
        cell.value = header
        cell.font = _FONT_SMALL_BOLD
        cell.fill = _GRAY_FILL
        cell.border = _THIN_BORDER
        cell.alignment = _WRAP_CENTER
    ws.row_dimensions[row].height = 28  # two-line header
    row += 1

    # Product rows
    for prod in data.products:
        ws.cell(row=row, column=1, value=prod.code).font = _FONT_SMALL
        ws.cell(row=row, column=2, value=prod.description).font = _FONT_SMALL
        ws.cell(row=row, column=3, value=_format_qty(prod.quantity)).font = _FONT_SMALL
        ws.cell(row=row, column=4, value=_format_price(prod.unit_price)).font = _FONT_SMALL
        ws.cell(row=row, column=5, value=_format_price(prod.total_price)).font = _FONT_SMALL

        ws.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="center")
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right", vertical="center")

        for ci in range(1, 6):
            ws.cell(row=row, column=ci).border = _THIN_BORDER

        # Row height based on content
        desc_lines = _count_lines(prod.description, 48)
        code_lines = _count_lines(prod.code, 12) if prod.code else 1
        ws.row_dimensions[row].height = max(_LINE_HEIGHT_SMALL, max(desc_lines, code_lines) * _LINE_HEIGHT_SMALL)
        row += 1

    # ── Summary rows ──

    # TOTAL IN SAR
    _add_summary_row(ws, row, "TOTAL IN SAR", _format_price(data.subtotal))
    row += 1

    # Installation Services (option 2/3)
    if has_installation:
        rate = data.installation_amount / data.device_count if data.device_count else 0
        ws.cell(row=row, column=2, value="Installation Services").font = _FONT_SMALL_BOLD
        ws.cell(row=row, column=3, value=str(data.device_count)).font = _FONT_SMALL
        ws.cell(row=row, column=3).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=4, value=_format_price(rate)).font = _FONT_SMALL
        ws.cell(row=row, column=4).alignment = Alignment(horizontal="right", vertical="center")
        ws.cell(row=row, column=5, value=_format_price(data.installation_amount)).font = _FONT_SMALL_BOLD
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right", vertical="center")
        for ci in range(1, 6):
            ws.cell(row=row, column=ci).border = _THIN_BORDER
        row += 1

    # VAT
    _add_summary_row(ws, row, "VAT", _format_price(data.vat))
    row += 1

    # GRAND TOTAL IN SAR
    _add_summary_row(ws, row, "GRAND TOTAL IN SAR", _format_price(data.grand_total))
    row += 1

    return row


def _add_summary_row(ws, row: int, label: str, value: str) -> None:
    """TOTAL / VAT / GRAND TOTAL row with borders on all 5 cells."""
    ws.cell(row=row, column=2, value=label).font = _FONT_SMALL_BOLD
    ws.cell(row=row, column=5, value=value).font = _FONT_SMALL_BOLD
    ws.cell(row=row, column=5).alignment = Alignment(horizontal="right", vertical="center")
    for ci in range(1, 6):
        ws.cell(row=row, column=ci).border = _THIN_BORDER


# ═══════════════════════════════════════════════════════════════
# Image helpers
# ════════════════���══════════════════════════════════════════════

def _add_image_from_bytes(ws, row: int, image_bytes: bytes, width_cm: float = 3.5) -> int:
    """Embed an image (signature etc.) from bytes."""
    img = XlImage(io.BytesIO(image_bytes))
    target_w = int(width_cm * 37.8)  # cm → px at 96 dpi
    if img.width and img.height:
        aspect = img.height / img.width
        img.width = target_w
        img.height = int(target_w * aspect)
    ws.add_image(img, f"A{row}")
    row += 3
    return row
